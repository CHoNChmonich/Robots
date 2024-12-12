import datetime
import json

import openpyxl
from django.db.models import Count
from openpyxl.utils import get_column_letter
from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.utils.timezone import now

from .models import Robot
from .validators import validate_robot_data

@csrf_exempt
def robot_create_post_view(request):
    if request.method == 'POST':
        if not request.body:
            return JsonResponse({"error": "No data provided"}, status=400)
        try:
            data = json.loads(request.body)
            validated_data = validate_robot_data(data)
            robot = Robot.objects.create(**validated_data)
            return JsonResponse({"message": "Robot created successfully", "robot_id": robot.id}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)

        except ValidationError as e:
            return JsonResponse({"error": str(e)}, status=400)

        except Exception as e:
            return JsonResponse({"error": f"Unexpected error: {str(e)}"}, status=500)


def robot_report_view(request):
    # Просто отображаем шаблон с кнопкой для генерации отчета
    return render(request, 'robots/robot_report.html')


def generate_excel_report(request):
    # Определяем временные рамки текущей недели
    today = now().date()
    start_of_week = today - datetime.timedelta(days=today.weekday())
    end_of_week = start_of_week + datetime.timedelta(days=6)

    # Фильтруем роботов, созданных на текущей неделе
    robots_this_week = Robot.objects.filter(created__date__gte=start_of_week, created__date__lte=end_of_week)

    # Группируем по модели и версии и считаем количество
    grouped_robots = robots_this_week.values('model', 'version').annotate(total=Count('id'))

    # Создаем Excel-файл
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Отчет по роботам"

    # Заголовки для таблицы
    headers = ['Модель', 'Версия', 'Количество']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Заполняем данные
    for row_num, robot in enumerate(grouped_robots, start=2):
        sheet[f"A{row_num}"] = robot['model']
        sheet[f"B{row_num}"] = robot['version']
        sheet[f"C{row_num}"] = robot['total']

    # Сохраняем Excel в ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=robots_report_{start_of_week}.xlsx'
    workbook.save(response)
    return response